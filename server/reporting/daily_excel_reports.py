from __future__ import annotations

import asyncio
import json
import os
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session, joinedload

from models import Card, Channel


DATE_FMT = "%d-%b-%Y"


HEADER_ALIASES: Dict[str, List[str]] = {
    "Work Order No.": ["W.O.NO.", "W.O.NO", "WO", "WONO", "WO NO", "WORK ORDER NO", "WORKORDERNO"],
    "Customer": ["CUSTOMER", "CUSTOMER NAME"],
    "Tank Size": ["TANK SIZE", "TANKSIZE"],
    "Brand": ["BRAND"],
    "Type (INS / NON)": ["TYPE", "TYPE (INS / NON)", "INS / NON"],
    "Type": ["TYPE", "TYPE (INS / NON)", "INS / NON"],
    "Location": ["LOCATION", "PROJECT LOCATION"],
    "Delivery Date": ["DEL.DATE", "DEL DATE", "DELIVERY DATE", "DEL.DATE."],
    "Delivery Status": ["DEL.STATUS", "DEL STATUS", "DELIVERY STATUS", "DEL.STATUS."],
    "Delivery Completion Date": ["DEL.COMP.DATE", "DEL.COMP. DATE", "DEL COMP DATE", "DELIVERY COMPLETION DATE"],
    "Contact Person": ["CONT.PER.", "CONT.PER", "CONTACT PERSON", "CONTACT PERSON (CUSTOMER NAME)", "CONT PER"],
    "Phone Number": ["PH.NUMBER", "PH. NUMBER", "PHONE NUMBER", "PHONE", "PH NUMBER"],
    "Sales Person": ["SALES.PER.", "SALES.PER", "SALES PERSON", "SALESPERSON"],
    "Remarks": ["REMARKS", "REMARK"],
    "Installation Start Date": ["INST.START", "INST.START.", "INSTALLATION START DATE", "INST START"],
    "Installation Status": ["INS.STATUS", "INST.STATUS", "STATUS OF INSTALATION", "STATUS OF INSTALLATION", "INSTALLATION STATUS"],
    "Workers": ["WORKERS", "WORKER"],
    "Installation Completion Date": ["INS.COMPLET.", "INS.COMPL.DATE", "INS COMPLETION DATE", "INSTALLATION COMPLETION DATE"],
    "Installation Date": ["INSTALLATION DATE", "INSTA.", "INS.DATE", "INS DATE"],
    "Payment Status": ["PAYMENT STATUS", "PAY.STATUS", "PAY STATUS"],
    "PDC Cheque": ["PDC CHEQUE", "P.D.C. CHEUQE", "P.D.C. CHEQUE", "PDC CHECK"],
}


@dataclass
class ReportRecord:
    source_card_id: str
    work_order_no: str
    quotation_no: str
    customer: str
    tank_size: str
    brand: str
    tank_type: str
    location: str
    delivery_date: str
    installation_date: str
    delivery_status: str
    installation_status: str
    delivery_completion_date: str
    installation_completion_date: str
    contact_person: str
    phone_number: str
    sales_person: str
    workers: str
    payment_percent: int
    payment_status: str
    pdc_cheque: str
    remarks: str
    schedule_type: str
    schedule_stage: str
    schedule_list_id: str
    requires_installation: Optional[bool]
    completed_date: Optional[date]
    updated_date: Optional[date]


@dataclass
class ScheduleOverlay:
    source_card_id: str
    workers: str
    brand: str
    delivery_status: str
    installation_status: str


@dataclass
class ScheduleCardSnapshot:
    source_card_id: str
    wo_code: str
    list_id: str
    schedule_type: str
    is_confirmed: bool
    confirmed_date: Optional[date]
    completed_date: Optional[date]
    workers: str
    brand: str
    delivery_status: str
    installation_status: str
    customer: str
    tank_size: str
    location: str
    contact_person: str
    phone_number: str
    sales_person: str
    payment_percent: int
    latest_remark_text: str
    latest_remark_at: Optional[date]


@dataclass(frozen=True)
class ReportSpec:
    key: str
    template_name: str
    output_name: str
    columns: List[str]
    include: Callable[[ReportRecord, date], bool]


def _today_local() -> date:
    return datetime.now().date()


def _as_date(value: Optional[datetime]) -> Optional[date]:
    if not value:
        return None
    return value.date()


def _fmt_date(value: Optional[date]) -> str:
    if not value:
        return ""
    return value.strftime(DATE_FMT)


def _normalize_text(value: str) -> str:
    return "".join(ch for ch in (value or "").lower() if ch.isalnum())


def _infer_schedule_type(card: Card) -> str:
    if card.schedule_type in {"Delivery", "Installation"}:
        return card.schedule_type
    stage = (card.schedule_stage or "").lower()
    if "installation" in stage:
        return "Installation"
    if "delivery" in stage:
        return "Delivery"
    if card.list_rel and card.list_rel.list_name in {"Installation", "Delivery"}:
        return card.list_rel.list_name
    return "Delivery"


def _latest_remark(card: Card) -> str:
    if not card.remarks:
        return ""
    sorted_remarks = sorted(
        card.remarks,
        key=lambda r: (r.updated_at or r.created_at or datetime.min),
        reverse=True,
    )
    return (sorted_remarks[0].description or "").strip()


def _tank_size(card: Card) -> str:
    wo = card.work_order_details
    if not wo or not wo.items:
        return ""
    for item in wo.items:
        desc = (item or {}).get("itemDescription", "")
        if isinstance(desc, str) and desc.strip():
            return desc.strip()
    return ""


def _tank_type(card: Card) -> str:
    wo = card.work_order_details
    if not wo:
        return ""
    insulated = bool(wo.type_insulated)
    non_insulated = bool(wo.type_non_insulated)
    if insulated and non_insulated:
        return "INS / NON"
    if insulated:
        return "INS"
    if non_insulated:
        return "NON"
    return ""


def _contact_person(card: Card) -> str:
    wo = card.work_order_details
    # Business expectation from template: CONT.PER should carry customer name.
    if card.customer_name and card.customer_name.strip():
        return card.customer_name.strip()
    if wo and wo.delivery_contact_name and wo.delivery_contact_name.strip():
        return wo.delivery_contact_name.strip()
    if wo and wo.company_contact_name and wo.company_contact_name.strip():
        return wo.company_contact_name.strip()
    return ""


def _phone(card: Card) -> str:
    wo = card.work_order_details
    if not wo:
        return ""
    return (wo.delivery_contact_number or wo.company_phone or "").strip()


def _delivery_completed_on(card: Card) -> Optional[date]:
    stage = (card.schedule_stage or "").lower()
    if "delivery completed" in stage:
        return _as_date(card.completed_at)
    return None


def _installation_completed_on(card: Card) -> Optional[date]:
    stage = (card.schedule_stage or "").lower()
    if "installation completed" in stage:
        return _as_date(card.completed_at)
    return None


def _payment_status(payment_percent: int) -> str:
    return f"{max(0, min(100, int(payment_percent)))}%"


def _parse_percent_number(value: Optional[str]) -> float:
    if not value:
        return 0.0
    cleaned = "".join(ch for ch in str(value) if (ch.isdigit() or ch in ".-"))
    if not cleaned:
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _pdc_cheque_percent(card: Card) -> str:
    oc = card.order_confirmation
    if not oc:
        return "0%"
    total = 0.0
    if oc.advance_pdc:
        total += _parse_percent_number(oc.advance_percent)
    if oc.delivery_pdc:
        total += _parse_percent_number(oc.delivery_percent)
    if oc.completion_pdc:
        total += _parse_percent_number(oc.completion_amount)
    if oc.testing_commissioning_pdc:
        total += _parse_percent_number(oc.testing_commissioning_amount)
    if oc.retention_pdc:
        total += _parse_percent_number(oc.retention_amount)

    if abs(total - round(total)) < 1e-9:
        return f"{int(round(total))}%"
    return f"{round(total, 2)}%"


def _normalize_overlay_brand(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    upper = raw.upper()
    if "COLEX" in upper:
        return "COLEX TANKS"
    if "PIPECO" in upper or "PIPECO" in upper:
        return "PIPECO TANKS"
    return raw


def _payment_completed_on(rec: ReportRecord) -> Optional[date]:
    if rec.payment_percent < 100:
        return None
    return rec.updated_date or rec.completed_date


def _completed_before_today(done_on: Optional[date], today: date) -> bool:
    return bool(done_on and done_on < today)


def _delivery_scope(rec: ReportRecord) -> bool:
    return rec.schedule_type == "Delivery" or "delivery" in rec.schedule_stage.lower()


def _installation_scope(rec: ReportRecord) -> bool:
    return rec.schedule_type == "Installation" or "installation" in rec.schedule_stage.lower()


def _delivery_planning(rec: ReportRecord, today: date) -> bool:
    done_on = _parse_report_date(rec.delivery_completion_date)
    return _delivery_scope(rec) and not _completed_before_today(done_on, today)


def _delivery_pending(rec: ReportRecord, today: date) -> bool:
    done_on = _parse_report_date(rec.delivery_completion_date)
    if not _delivery_scope(rec):
        return False
    if done_on:
        return False
    return "pending" in rec.delivery_status.lower() or "scheduled" in rec.delivery_status.lower() or rec.delivery_status == ""


def _delivery_status(rec: ReportRecord, today: date) -> bool:
    if not _delivery_scope(rec):
        return False
    done_on = _parse_report_date(rec.delivery_completion_date)
    if not done_on:
        return True
    return done_on == today


def _installation_planning(rec: ReportRecord, today: date) -> bool:
    done_on = _parse_report_date(rec.installation_completion_date)
    return _installation_scope(rec) and not _completed_before_today(done_on, today)


def _installation_pending(rec: ReportRecord, today: date) -> bool:
    done_on = _parse_report_date(rec.installation_completion_date)
    if not _installation_scope(rec):
        return False
    if done_on:
        return False
    return "pending" in rec.installation_status.lower() or "scheduled" in rec.installation_status.lower() or rec.installation_status == ""


def _installation_status(rec: ReportRecord, today: date) -> bool:
    if not _installation_scope(rec):
        return False
    done_on = _parse_report_date(rec.installation_completion_date)
    if not done_on:
        return True
    return done_on == today


def _payment_report(rec: ReportRecord, today: date) -> bool:
    if rec.payment_percent < 100:
        return True
    paid_on = _payment_completed_on(rec)
    if not paid_on:
        return False
    # Keep card on payment sheet for one additional day after it reaches 100%.
    return today <= (paid_on + timedelta(days=1))


def _parse_report_date(value: str) -> Optional[date]:
    if not value:
        return None
    try:
        return datetime.strptime(value, DATE_FMT).date()
    except ValueError:
        return None


def _to_record(card: Card) -> ReportRecord:
    wo = card.work_order_details
    delivery_done = _delivery_completed_on(card)
    installation_done = _installation_completed_on(card)
    payment_percent = int(card.payment_percent or 0)
    schedule_type = _infer_schedule_type(card)
    stage_text = (card.schedule_stage or "").strip()
    status_fallback = ""
    if card.user_work_status is not None:
        status_fallback = getattr(card.user_work_status, "value", str(card.user_work_status))
    if not status_fallback:
        status_fallback = (card.list_rel.list_name if card.list_rel else "") or "Pending"

    delivery_status = ""
    installation_status = ""
    if schedule_type == "Delivery":
        delivery_status = stage_text or status_fallback
    elif schedule_type == "Installation":
        installation_status = stage_text or status_fallback

    return ReportRecord(
        source_card_id=str(card.id),
        work_order_no=(card.work_order_number or "").strip(),
        quotation_no=(card.quote_number or "").strip(),
        customer=(card.customer_company_name or (wo.company_name if wo else "") or card.customer_name or "").strip(),
        tank_size=_tank_size(card),
        brand=(wo.brand.value if wo and wo.brand else "").strip(),
        tank_type=_tank_type(card),
        location=(card.project_location or (wo.delivery_location if wo else "") or "").strip(),
        delivery_date=_fmt_date((wo.delivery_date if wo and wo.delivery_date else card.date)),
        installation_date=_fmt_date((wo.installation_completion_date if wo and wo.installation_completion_date else card.date)),
        delivery_status=delivery_status,
        installation_status=installation_status,
        delivery_completion_date=_fmt_date(delivery_done),
        installation_completion_date=_fmt_date(installation_done),
        contact_person=_contact_person(card),
        phone_number=_phone(card),
        sales_person=(card.sales_person or "").strip(),
        workers="",
        payment_percent=payment_percent,
        payment_status=_payment_status(payment_percent),
        pdc_cheque=_pdc_cheque_percent(card),
        remarks=_latest_remark(card),
        schedule_type=schedule_type,
        schedule_stage=stage_text,
        schedule_list_id="",
        requires_installation=(bool(wo.installation) if wo else None),
        completed_date=_as_date(card.completed_at),
        updated_date=_as_date(card.updated_at),
    )


def _column_value(rec: ReportRecord, col: str) -> str:
    installation_for_payment = rec.installation_completion_date or "NO"
    mapping: Dict[str, str] = {
        "Work Order No.": rec.work_order_no,
        "W.O.NO.": rec.work_order_no,
        "WO": rec.work_order_no,
        "Customer": rec.customer,
        "Tank Size": rec.tank_size,
        "Brand": rec.brand,
        "Type (INS / NON)": rec.tank_type,
        "Type": rec.tank_type,
        "Location": rec.location,
        "Delivery Date": rec.delivery_date,
        "DEL.DATE": rec.delivery_date,
        "Delivery Status": rec.delivery_status,
        "DEL.STATUS": rec.delivery_status,
        "Delivery Completion Date": rec.delivery_completion_date,
        "DEL.COMP.DATE": rec.delivery_completion_date,
        "Contact Person": rec.contact_person,
        "CONT.PER.": rec.contact_person,
        "Phone Number": rec.phone_number,
        "PH.NUMBER": rec.phone_number,
        "Sales Person": rec.sales_person,
        "SALES.PER.": rec.sales_person,
        "Remarks": rec.remarks,
        "Installation Start Date": rec.installation_date,
        "INST.START": rec.installation_date,
        "Installation Status": rec.installation_status,
        "INS.STATUS": rec.installation_status,
        "Workers": rec.workers,
        "Installation Completion Date": rec.installation_completion_date,
        "INS.COMPLET.": rec.installation_completion_date,
        "Installation Date": installation_for_payment,
        "INSTA.": installation_for_payment,
        "Payment Status": rec.payment_status,
        "PAY.STATUS": rec.payment_status,
        "PDC Cheque": rec.pdc_cheque,
        "P.D.C. CHEUQE": rec.pdc_cheque,
    }
    return mapping.get(col, "")


def _load_schedule_overlays() -> Dict[str, ScheduleOverlay]:
    path = Path(os.getenv("REPORT_SCHEDULE_DATA_FILE", "/app/client_data/schedule-data.json"))
    if not path.exists():
        return {}

    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}

    overlays: Dict[str, ScheduleOverlay] = {}
    if not isinstance(raw, dict):
        return overlays

    for cards in raw.values():
        if not isinstance(cards, list):
            continue
        for row in cards:
            if not isinstance(row, dict):
                continue
            src_id = str(row.get("sourceCardId") or "").strip()
            if not src_id:
                continue
            workers_raw = row.get("workers")
            workers = ", ".join([str(w).strip() for w in workers_raw if str(w).strip()]) if isinstance(workers_raw, list) else ""
            overlays[src_id] = ScheduleOverlay(
                source_card_id=src_id,
                workers=workers,
                brand=_normalize_overlay_brand(str(row.get("brand") or "")),
                delivery_status=str(row.get("deliveryStatus") or "").strip(),
                installation_status=str(row.get("installationStatus") or "").strip(),
            )
    return overlays


def _parse_iso_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        if "T" in text:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
        return date.fromisoformat(text)
    except Exception:
        return None


def _as_int(value: Any, fallback: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return fallback


def _load_schedule_cards() -> List[ScheduleCardSnapshot]:
    path = Path(os.getenv("REPORT_SCHEDULE_DATA_FILE", "/app/client_data/schedule-data.json"))
    if not path.exists():
        return []

    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []

    if not isinstance(raw, dict):
        return []

    cards: List[ScheduleCardSnapshot] = []

    def latest_schedule_remark(row: Dict[str, Any]) -> tuple[str, Optional[date]]:
        remarks = row.get("remarks")
        if not isinstance(remarks, list) or not remarks:
            return "", None

        best_text = ""
        best_at: Optional[date] = None
        best_dt: Optional[datetime] = None

        for item in remarks:
            if not isinstance(item, dict):
                continue
            text = str(item.get("text") or "").strip()
            author = str(item.get("author") or "").strip()
            at_raw = item.get("at")
            at_dt: Optional[datetime] = None
            if at_raw:
                try:
                    at_dt = datetime.fromisoformat(str(at_raw).replace("Z", "+00:00"))
                except Exception:
                    at_dt = None

            if not text and not author:
                continue

            if best_dt is None or (at_dt is not None and at_dt > best_dt):
                best_dt = at_dt
                best_at = at_dt.date() if at_dt else None
                best_text = f"{author}: {text}".strip(": ")

        return best_text, best_at

    for list_id, rows in raw.items():
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            workers_raw = row.get("workers")
            workers = ", ".join([str(w).strip() for w in workers_raw if str(w).strip()]) if isinstance(workers_raw, list) else ""
            latest_text, latest_at = latest_schedule_remark(row)
            cards.append(
                ScheduleCardSnapshot(
                    source_card_id=str(row.get("sourceCardId") or "").strip(),
                    wo_code=str(row.get("woCode") or "").strip(),
                    list_id=str(row.get("listId") or list_id or "").strip(),
                    schedule_type=str(row.get("scheduleType") or "").strip() or ("Installation" if str(list_id).startswith("installation-") else "Delivery"),
                    is_confirmed=bool(row.get("isConfirmed")),
                    confirmed_date=_parse_iso_date(row.get("confirmedDate")),
                    completed_date=_parse_iso_date(row.get("completedDate")),
                    workers=workers,
                    brand=_normalize_overlay_brand(str(row.get("brand") or "")),
                    delivery_status=str(row.get("deliveryStatus") or "").strip(),
                    installation_status=str(row.get("installationStatus") or "").strip(),
                    customer=str(row.get("customer") or "").strip(),
                    tank_size=str(row.get("tankSize") or "").strip(),
                    location=str(row.get("location") or "").strip(),
                    contact_person=str(row.get("contactPerson") or "").strip(),
                    phone_number=str(row.get("phone") or "").strip(),
                    sales_person=str(row.get("salesPerson") or "").strip(),
                    payment_percent=max(0, min(100, _as_int(row.get("paymentPercent"), 0))),
                    latest_remark_text=latest_text,
                    latest_remark_at=latest_at,
                )
            )
    return cards


def _record_from_schedule_card(base: Optional[ReportRecord], sc: ScheduleCardSnapshot, today: date) -> ReportRecord:
    if base:
        rec = ReportRecord(**{**base.__dict__})
    else:
        rec = ReportRecord(
            source_card_id=sc.source_card_id,
            work_order_no=sc.wo_code,
            quotation_no="",
            customer=sc.customer,
            tank_size=sc.tank_size,
            brand=sc.brand,
            tank_type="",
            location=sc.location,
            delivery_date="",
            installation_date="",
            delivery_status="",
            installation_status="",
            delivery_completion_date="",
            installation_completion_date="",
            contact_person=sc.contact_person,
            phone_number=sc.phone_number,
            sales_person=sc.sales_person,
            workers="",
            payment_percent=sc.payment_percent,
            payment_status=_payment_status(sc.payment_percent),
            pdc_cheque="0%",
            remarks="",
            schedule_type=sc.schedule_type,
            schedule_stage="",
            schedule_list_id=sc.list_id,
            requires_installation=None,
            completed_date=sc.completed_date,
            updated_date=today,
        )

    rec.schedule_list_id = sc.list_id
    if sc.workers:
        rec.workers = sc.workers
    if (not rec.brand.strip()) and sc.brand:
        rec.brand = sc.brand
    if sc.delivery_status:
        rec.delivery_status = sc.delivery_status
    if sc.installation_status:
        rec.installation_status = sc.installation_status
    if (not rec.work_order_no.strip()) and sc.wo_code:
        rec.work_order_no = sc.wo_code
    if (not rec.customer.strip()) and sc.customer:
        rec.customer = sc.customer
    if (not rec.tank_size.strip()) and sc.tank_size:
        rec.tank_size = sc.tank_size
    if (not rec.location.strip()) and sc.location:
        rec.location = sc.location
    if (not rec.contact_person.strip()) and sc.contact_person:
        rec.contact_person = sc.contact_person
    if (not rec.phone_number.strip()) and sc.phone_number:
        rec.phone_number = sc.phone_number
    if (not rec.sales_person.strip()) and sc.sales_person:
        rec.sales_person = sc.sales_person

    rec.payment_percent = max(0, min(100, sc.payment_percent if sc.payment_percent or not base else rec.payment_percent))
    rec.payment_status = _payment_status(rec.payment_percent)
    if sc.latest_remark_text:
        rec.remarks = sc.latest_remark_text

    if sc.list_id.startswith("delivery-"):
        day_part = sc.list_id.replace("delivery-", "", 1)
        try:
            list_day = date.fromisoformat(day_part)
            rec.delivery_date = _fmt_date(list_day)
        except Exception:
            pass
    if sc.list_id.startswith("installation-"):
        day_part = sc.list_id.replace("installation-", "", 1)
        try:
            list_day = date.fromisoformat(day_part)
            rec.installation_date = _fmt_date(list_day)
        except Exception:
            pass

    if sc.list_id.startswith("delivery-") and sc.is_confirmed:
        done_on = sc.confirmed_date or today
        rec.delivery_completion_date = _fmt_date(done_on)
        if not rec.delivery_status:
            rec.delivery_status = "Delivery completed"

    if sc.list_id.startswith("installation-") and sc.completed_date:
        rec.installation_completion_date = _fmt_date(sc.completed_date)
        if not rec.installation_status:
            rec.installation_status = "Installation completed"

    if not rec.schedule_type:
        rec.schedule_type = sc.schedule_type
    return rec


def _build_records_by_schedule(
    db_records: List[ReportRecord],
    schedule_cards: List[ScheduleCardSnapshot],
    today: date,
) -> Dict[str, List[ReportRecord]]:
    tomorrow = today + timedelta(days=1)
    d_tomorrow_key = f"delivery-{tomorrow.isoformat()}"
    i_tomorrow_key = f"installation-{tomorrow.isoformat()}"
    d_today_key = f"delivery-{today.isoformat()}"
    i_today_key = f"installation-{today.isoformat()}"

    by_source = {rec.source_card_id: rec for rec in db_records if rec.source_card_id}
    by_wo = {rec.work_order_no: rec for rec in db_records if rec.work_order_no}

    def materialize(sc: ScheduleCardSnapshot) -> ReportRecord:
        base = by_source.get(sc.source_card_id) if sc.source_card_id else None
        if not base and sc.wo_code:
            base = by_wo.get(sc.wo_code)
        return _record_from_schedule_card(base, sc, today)

    buckets: Dict[str, List[ReportRecord]] = {
        "delivery_pending": [],
        "delivery_planning": [],
        "delivery_status": [],
        "installation_pending": [],
        "installation_planning": [],
        "installation_status": [],
        "payment_status": [],
    }

    for sc in schedule_cards:
        lid = sc.list_id
        if lid == "pending-delivery":
            buckets["delivery_pending"].append(materialize(sc))
        elif lid == d_tomorrow_key:
            buckets["delivery_planning"].append(materialize(sc))
        elif lid == "pending-installation":
            buckets["installation_pending"].append(materialize(sc))
        elif lid == i_tomorrow_key:
            buckets["installation_planning"].append(materialize(sc))

    buckets["delivery_status"] = (
        list(buckets["delivery_pending"]) +
        list(buckets["delivery_planning"]) +
        [
            rec
            for sc in schedule_cards
            if sc.list_id == d_today_key and sc.is_confirmed
            for rec in [materialize(sc)]
            if rec.requires_installation is not False
        ]
    )

    buckets["installation_status"] = (
        list(buckets["installation_pending"]) +
        list(buckets["installation_planning"]) +
        [
            materialize(sc)
            for sc in schedule_cards
            if sc.list_id == i_today_key and sc.completed_date is not None
        ]
    )

    # Payment Status:
    # 1) Installation-completed cards are tracked daily until payment reaches 100%.
    # 2) Delivery-only cards that are delivered are also tracked daily until 100%.
    seen_keys = set()
    for sc in schedule_cards:
        is_delivery_done = sc.list_id.startswith("delivery-") and sc.is_confirmed
        is_installation_done = sc.list_id.startswith("installation-") and sc.completed_date is not None
        if not (is_delivery_done or is_installation_done):
            continue

        rec = materialize(sc)
        key = rec.source_card_id or rec.work_order_no
        if not key:
            continue

        if rec.requires_installation is True:
            eligible = is_installation_done
        elif rec.requires_installation is False:
            eligible = is_delivery_done
        else:
            # Fallback for schedule-only rows without WO metadata.
            eligible = is_installation_done or is_delivery_done

        if not eligible:
            continue
        if not _payment_report(rec, today):
            continue
        if key in seen_keys:
            continue
        seen_keys.add(key)
        buckets["payment_status"].append(rec)

    return buckets


def _apply_schedule_overlays(records: List[ReportRecord], overlays: Dict[str, ScheduleOverlay], cards: List[Card]) -> List[ReportRecord]:
    card_by_work_order = {str(c.work_order_number or "").strip(): c for c in cards}
    updated: List[ReportRecord] = []

    for rec in records:
        card = card_by_work_order.get(rec.work_order_no)
        if not card:
            updated.append(rec)
            continue

        ov = overlays.get(str(card.id))
        if not ov:
            updated.append(rec)
            continue

        next_rec = ReportRecord(**{**rec.__dict__})
        if ov.workers:
            next_rec.workers = ov.workers
        if (not next_rec.brand.strip()) and ov.brand:
            next_rec.brand = ov.brand
        if ov.delivery_status:
            next_rec.delivery_status = ov.delivery_status
        if ov.installation_status:
            next_rec.installation_status = ov.installation_status
        updated.append(next_rec)

    return updated


def _find_header_row(ws, columns: Iterable[str]) -> tuple[int, Dict[str, int]]:
    expected: Dict[str, str] = {}
    for col in columns:
        expected[_normalize_text(col)] = col
        for alias in HEADER_ALIASES.get(col, []):
            expected[_normalize_text(alias)] = col
    best_row = 0
    best_map: Dict[str, int] = {}
    for row_idx in range(1, min(ws.max_row, 100) + 1):
        row_map: Dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if not isinstance(val, str):
                continue
            norm = _normalize_text(val)
            if norm in expected:
                row_map[expected[norm]] = col_idx
        if len(row_map) > len(best_map):
            best_map = row_map
            best_row = row_idx
        if len(row_map) == len(expected):
            break

    if len(best_map) < max(1, len(list(columns)) // 2):
        raise ValueError(f"Unable to find header row in worksheet '{ws.title}'")
    return best_row, best_map


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.protection = copy(src.protection)


def _write_rows_to_template(template_path: Path, output_path: Path, rows: List[ReportRecord], spec: ReportSpec) -> Path:
    wb = load_workbook(template_path)
    ws = wb.active
    header_row, col_map = _find_header_row(ws, spec.columns)
    data_row = header_row + 1
    max_col = max(col_map.values())

    for row_idx in range(data_row, ws.max_row + 1):
        for col in col_map.values():
            ws.cell(row=row_idx, column=col).value = None

    for i, rec in enumerate(rows):
        target_row = data_row + i
        if target_row > ws.max_row:
            ws.insert_rows(target_row)
        _copy_row_style(ws, data_row, target_row, max_col)
        for col_name, col_idx in col_map.items():
            ws.cell(row=target_row, column=col_idx).value = _column_value(rec, col_name)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        # If the target workbook is currently open in Excel, save to a
        # timestamped fallback file instead of failing the whole report run.
        fallback = output_path.with_name(
            f"{output_path.stem} ({datetime.now().strftime('%H%M%S')}){output_path.suffix}"
        )
        wb.save(fallback)
        return fallback


REPORT_SPECS: List[ReportSpec] = [
    ReportSpec(
        key="delivery_planning",
        template_name="Delivery Planning.xlsx",
        output_name="Delivery Planning",
        columns=[
            "Work Order No.", "Customer", "Tank Size", "Brand", "Type (INS / NON)", "Location",
            "Delivery Date", "Delivery Status", "Delivery Completion Date", "Contact Person",
            "Phone Number", "Sales Person", "Remarks",
        ],
        include=_delivery_planning,
    ),
    ReportSpec(
        key="delivery_pending",
        template_name="Delivery Pending.xlsx",
        output_name="Delivery Pending",
        columns=[
            "Work Order No.", "Customer", "Tank Size", "Brand", "Type", "Location", "Delivery Date",
            "Delivery Status", "Delivery Completion Date", "Contact Person", "Phone Number", "Sales Person", "Remarks",
        ],
        include=_delivery_pending,
    ),
    ReportSpec(
        key="delivery_status",
        template_name="Delivery Status.xlsx",
        output_name="Delivery Status",
        columns=[
            "Work Order No.", "Customer", "Tank Size", "Brand", "Type", "Location", "Delivery Date",
            "Delivery Status", "Delivery Completion Date", "Contact Person", "Phone Number", "Sales Person", "Remarks",
        ],
        include=_delivery_status,
    ),
    ReportSpec(
        key="installation_planning",
        template_name="Installation Planning.xlsx",
        output_name="Installation Planning",
        columns=[
            "Work Order No.", "Customer", "Tank Size", "Brand", "Type", "Location", "Installation Start Date",
            "Installation Status", "Workers", "Contact Person", "Phone Number", "Installation Completion Date",
            "Sales Person", "Remarks",
        ],
        include=_installation_planning,
    ),
    ReportSpec(
        key="installation_pending",
        template_name="Installation Pending.xlsx",
        output_name="Installation Pending",
        columns=[
            "Work Order No.", "Customer", "Tank Size", "Brand", "Type", "Location", "Installation Start Date",
            "Installation Status", "Workers", "Contact Person", "Phone Number", "Installation Completion Date",
            "Sales Person", "Remarks",
        ],
        include=_installation_pending,
    ),
    ReportSpec(
        key="installation_status",
        template_name="Installation Status.xlsx",
        output_name="Installation Status",
        columns=[
            "Work Order No.", "Customer", "Tank Size", "Brand", "Type", "Location", "Installation Start Date",
            "Installation Status", "Workers", "Contact Person", "Phone Number", "Installation Completion Date",
            "Sales Person", "Remarks",
        ],
        include=_installation_status,
    ),
    ReportSpec(
        key="payment_status",
        template_name="Payment Status.xlsx",
        output_name="Payment Status",
        columns=[
            "Work Order No.", "Customer", "Delivery Date", "Installation Date", "Payment Status", "PDC Cheque",
            "Sales Person", "Remarks",
        ],
        include=_payment_report,
    ),
]


def _query_work_order_cards(db: Session) -> List[Card]:
    channel = db.query(Channel).filter(Channel.channel_name == "Work Order").first()
    if not channel:
        return []

    cards = (
        db.query(Card)
        .options(
            joinedload(Card.work_order_details),
            joinedload(Card.order_confirmation),
            joinedload(Card.remarks),
            joinedload(Card.assigned_user),
            joinedload(Card.list_rel),
        )
        .filter(Card.channel_id == channel.channel_id)
        .all()
    )
    return cards


def _build_report_records(db: Session) -> List[ReportRecord]:
    cards = _query_work_order_cards(db)
    records = [_to_record(card) for card in cards]
    return _apply_schedule_overlays(records, _load_schedule_overlays(), cards)


def _get_report_spec(spec_key: str) -> ReportSpec:
    for spec in REPORT_SPECS:
        if spec.key == spec_key:
            return spec
    raise ValueError(f"Unknown report key: {spec_key}")


def _serialize_rows(rows: List[ReportRecord], columns: List[str]) -> List[Dict[str, str]]:
    return [{col: _column_value(rec, col) for col in columns} for rec in rows]


def get_pending_report_details(db: Session, run_date: Optional[date] = None) -> Dict[str, object]:
    today = run_date or _today_local()
    records = _build_report_records(db)
    schedule_cards = _load_schedule_cards()
    buckets = _build_records_by_schedule(records, schedule_cards, today)

    delivery_spec = _get_report_spec("delivery_pending")
    installation_spec = _get_report_spec("installation_pending")

    delivery_rows = buckets["delivery_pending"]
    installation_rows = buckets["installation_pending"]

    return {
        "date": today.isoformat(),
        "delivery": {
            "columns": delivery_spec.columns,
            "rows": _serialize_rows(delivery_rows, delivery_spec.columns),
        },
        "installation": {
            "columns": installation_spec.columns,
            "rows": _serialize_rows(installation_rows, installation_spec.columns),
        },
    }


def generate_daily_reports(db: Session, run_date: Optional[date] = None) -> Dict[str, str]:
    today = run_date or _today_local()
    template_dir = Path(os.getenv("REPORT_TEMPLATES_DIR", "/app/report_templates"))
    output_dir = Path(os.getenv("REPORT_OUTPUT_DIR", "/app/reports/daily"))

    records = _build_report_records(db)
    schedule_cards = _load_schedule_cards()
    schedule_buckets = _build_records_by_schedule(records, schedule_cards, today)

    outputs: Dict[str, str] = {}
    missing_templates: List[str] = []

    for spec in REPORT_SPECS:
        template_path = template_dir / spec.template_name
        if not template_path.exists():
            missing_templates.append(spec.template_name)
            continue
        rows = schedule_buckets.get(spec.key)
        if rows is None:
            rows = [rec for rec in records if spec.include(rec, today)]
        filename = f"{spec.output_name} - {today.strftime('%Y-%m-%d')}.xlsx"
        out_path = output_dir / filename
        saved_path = _write_rows_to_template(template_path, out_path, rows, spec)
        outputs[spec.key] = str(saved_path)

    if missing_templates:
        outputs["missing_templates"] = ", ".join(missing_templates)
    return outputs


async def _report_scheduler_loop() -> None:
    run_time_str = os.getenv("REPORT_RUN_TIME", "00:05")
    try:
        hh, mm = [int(p) for p in run_time_str.split(":", 1)]
        run_time = time(hour=hh, minute=mm)
    except Exception:
        run_time = time(hour=0, minute=5)

    from database import SessionLocal

    if os.getenv("REPORT_GENERATE_ON_STARTUP", "0") == "1":
        try:
            with SessionLocal() as db:
                generate_daily_reports(db, _today_local())
        except Exception as exc:
            print(f"[daily-reports] startup generation failed: {exc}")

    while True:
        now = datetime.now()
        next_run = datetime.combine(now.date(), run_time)
        if now >= next_run:
            next_run = next_run + timedelta(days=1)
        await asyncio.sleep(max(1, int((next_run - now).total_seconds())))

        try:
            with SessionLocal() as db:
                generate_daily_reports(db, _today_local())
                print(f"[daily-reports] generated for {_today_local().isoformat()}")
        except Exception as exc:
            print(f"[daily-reports] generation failed: {exc}")


def start_daily_report_task() -> Optional[asyncio.Task]:
    if os.getenv("DAILY_REPORTS_ENABLED", "0") != "1":
        return None
    return asyncio.create_task(_report_scheduler_loop())
