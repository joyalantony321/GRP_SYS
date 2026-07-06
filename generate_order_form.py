"""
Generates Order Confirmation Form (COLEX TANKS TRADING L.L.C) as an Excel file.
Each input field, checkbox box, and underline from the PDF is represented as
a separate Excel cell with appropriate borders.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

wb = Workbook()
ws = wb.active
ws.title = "Order Confirmation Form"

# ─── Page setup ────────────────────────────────────────────────────────────────
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.paperSize   = ws.PAPERSIZE_A4
ws.page_margins.left   = 0.5
ws.page_margins.right  = 0.5
ws.page_margins.top    = 0.75
ws.page_margins.bottom = 0.75
ws.sheet_view.showGridLines = False
ws.print_area = "A1:T60"

# ─── Border helpers ────────────────────────────────────────────────────────────
th  = Side(style="thin")
med = Side(style="medium")
ns  = Side(style=None)

BX   = Border(top=th,  bottom=th,  left=th,  right=th)   # full box
UL   = Border(bottom=th)                                   # underline
NB   = Border()                                            # none
MED  = Border(top=med, bottom=med, left=med, right=med)   # medium box
MEDT = Border(top=med)                                     # medium top line
MEDB = Border(bottom=med)

# ─── Font helpers ──────────────────────────────────────────────────────────────
FB   = Font(name="Calibri", bold=True, size=10)
FN   = Font(name="Calibri", size=9)
FSB  = Font(name="Calibri", bold=True, size=9)
FT   = Font(name="Calibri", bold=True, size=12, underline="single")
FCO  = Font(name="Calibri", bold=True, size=15)
FAR  = Font(name="Calibri", bold=True, size=16)
FTRN = Font(name="Calibri", bold=True, size=11)
FHU  = Font(name="Calibri", bold=True, size=9,  underline="single")

# ─── Alignment helpers ─────────────────────────────────────────────────────────
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)
AR = Alignment(horizontal="right",  vertical="center", wrap_text=True)

# ─── Column widths ─────────────────────────────────────────────────────────────
# 20 columns  A … T
# A=1(L-margin) | B=11(main label) | C–E=5,5,5(input) |
# F=5(%CDC/YES label) | G=3(chk□) | H=5(%PDC/NO label) | I=3(chk□) |
# J=5(BEFORE/% label) | K=3(chk□) | L=5(AFTER label) | M=3(chk□) |
# N=1(sep) | O=11(right label) | P–S=5,5,5,5(right input) | T=1(R-margin)
COL_W = [
    1,            # 1  A
    11,           # 2  B
    5, 5, 5,      # 3-5  C D E
    5,            # 6  F
    3,            # 7  G  ← checkbox
    5,            # 8  H
    3,            # 9  I  ← checkbox
    5,            # 10 J
    3,            # 11 K  ← checkbox
    5,            # 12 L
    3,            # 13 M  ← checkbox
    1,            # 14 N  separator
    11,           # 15 O
    5, 5, 5, 5,   # 16-19  P Q R S
    1,            # 20 T
]
for idx, w in enumerate(COL_W, 1):
    ws.column_dimensions[get_column_letter(idx)].width = w

# ─── Helpers ───────────────────────────────────────────────────────────────────
def s(row, col, val="", font=None, align=None, brd=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font:  cell.font      = font
    if align: cell.alignment = align
    if brd:   cell.border    = brd
    if fill:  cell.fill      = fill
    return cell

def m(r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def rh(row, h):
    ws.row_dimensions[row].height = h

def label(row, col_start, col_end, text, font=None, align=None, brd=None):
    """Merge cols and write a label."""
    if col_start != col_end:
        m(row, col_start, row, col_end)
    s(row, col_start, text, font or FSB, align or AL, brd)

def input_box(row, col_start, col_end, brd=BX):
    """Merge cols as an input field (box border)."""
    if col_start != col_end:
        m(row, col_start, row, col_end)
    s(row, col_start, "", FN, AL, brd)

def input_line(row, col_start, col_end):
    """Merge cols as an underlined input field."""
    if col_start != col_end:
        m(row, col_start, row, col_end)
    s(row, col_start, "", FN, AL, UL)

def checkbox(row, col, label_text=""):
    """Single cell with full-box border (the □)."""
    s(row, col, "", FN, AC, BX)

def chk_group(row, c_lbl1, c_chk1, lbl1, c_lbl2, c_chk2, lbl2):
    """Two label+checkbox pairs side by side."""
    s(row, c_lbl1, lbl1, FSB, AC)
    checkbox(row, c_chk1)
    s(row, c_lbl2, lbl2, FSB, AC)
    checkbox(row, c_chk2)

def full_label(row, text, font=None, align=None, brd=None):
    """Full-width label across body cols 2-20."""
    m(row, 2, row, 20)
    s(row, 2, text, font or FSB, align or AL, brd)

def section_header(row, text):
    """Underlined section header."""
    m(row, 2, row, 13)
    s(row, 2, text, FHU, AL)

# ═══════════════════════════════════════════════════════════════════════════════
#  ROW LAYOUT
# ═══════════════════════════════════════════════════════════════════════════════

# ── Row 1: top margin ──────────────────────────────────────────────────────────
rh(1, 4)

# ── Rows 2-4: Header (logo | Arabic | company | TRN) ─────────────────────────
rh(2, 30); rh(3, 22); rh(4, 16)

# Logo box  (A–B, rows 2-3)
m(2, 1, 3, 2)
s(2, 1, "LOGO", FSB, AC, BX)

# Arabic heading (C–N row 2)
m(2, 3, 2, 14)
s(2, 3, "كولـكس لتجـارة الـصهـاريـج ذ.م.م", FAR, AC)

# Company name (C–N row 3)
m(3, 3, 3, 14)
s(3, 3, "COLEX TANKS TRADING L.L.C", FCO, AC)

# TRN (full row 4)
m(4, 1, 4, 20)
s(4, 1, "TRN No.: 100508297700003", FTRN, AC)

# ── Row 5: heavy separator line ───────────────────────────────────────────────
rh(5, 3)
for c in range(1, 21):
    ws.cell(row=5, column=c).border = MEDT

# ── Row 6: Form title ─────────────────────────────────────────────────────────
rh(6, 22)
m(6, 1, 6, 20)
s(6, 1, "ORDER CONFIRMATION FORM", FT, AC)

# ── Row 7: spacer ─────────────────────────────────────────────────────────────
rh(7, 6)

# ── Row 8: WORK ORDER | LPO NO ───────────────────────────────────────────────
rh(8, 20)
s(8, 2, "WORK ORDER", FSB, AL)
input_box(8, 3, 9)                        # cols 3-9 → input box
s(8, 15, "LPO NO:", FSB, AL)
input_box(8, 16, 19)                      # cols 16-19 → input box

# ── Row 9: spacer ─────────────────────────────────────────────────────────────
rh(9, 6)

# ── Row 10: QTN NO | DATE ────────────────────────────────────────────────────
rh(10, 20)
s(10, 2, "QTN NO:", FSB, AL)
input_box(10, 3, 9)
s(10, 15, "DATE:", FSB, AL)
input_box(10, 16, 19)

# ── Row 11: spacer ────────────────────────────────────────────────────────────
rh(11, 6)

# ── Row 12: DID YOU CONFIRM… ─────────────────────────────────────────────────
rh(12, 14)
full_label(12, "DID YOU CONFIRM THE FOLLOWING TERMS IN THE LPO?", FSB)

# ── Row 13: spacer ────────────────────────────────────────────────────────────
rh(13, 6)

# ── Rows 14-16: YES/NO checkboxes for 3 items ────────────────────────────────
rh(14, 16); rh(15, 16); rh(16, 16)

# Row 14: TANK BRAND, SIZE, TYPE VALUE?
m(14, 3, 14, 5)
s(14, 3, "TANK BRAND, SIZE, TYPE VALUE?", FN, AL)
chk_group(14, 6, 7, "YES", 8, 9, "NO")

# Row 15: PAYMENT TERMS?
m(15, 3, 15, 5)
s(15, 3, "PAYMENT TERMS?", FN, AL)
chk_group(15, 6, 7, "YES", 8, 9, "NO")

# Row 16: OTHER TERMS & CONDITION?
m(16, 3, 16, 5)
s(16, 3, "OTHER TERMS & CONDITION?", FN, AL)
chk_group(16, 6, 7, "YES", 8, 9, "NO")

# ── Row 17: MENTION IF THERE ANY PENALITY… ───────────────────────────────────
rh(17, 14)
m(17, 2, 17, 5)
s(17, 2, "MENTION IF THERE ANY PENALITY/CONDITIONS/TIME PERIOD IN THE LPO", FN, AL)
input_line(17, 6, 19)

# ── Row 18: spacer ────────────────────────────────────────────────────────────
rh(18, 8)

# ── Row 19: DETAILS OF PAYMENT TERMS ─────────────────────────────────────────
rh(19, 16)
section_header(19, "DETAILS OF PAYMENT TERMS")

# ── Row 20: spacer ────────────────────────────────────────────────────────────
rh(20, 6)

# ── Row 21: ADVANCE ──────────────────────────────────────────────────────────
rh(21, 16)
s(21, 2, "ADVANCE", FSB, AL)
input_line(21, 3, 5)          # % value input
s(21, 6, "%CDC", FSB, AC)
checkbox(21, 7)
input_line(21, 8, 9)          # second % value
s(21, 10, "%PDC", FSB, AC)
checkbox(21, 11)

# ── Row 22: PAYMENT COLLECTION ───────────────────────────────────────────────
rh(22, 16)
s(22, 2, "PAYMENT COLLECTION", FSB, AL)
s(22, 5, "FROM SITE", FSB, AC)
checkbox(22, 7)
s(22, 8, "OR OFFICE", FSB, AC)
checkbox(22, 11)

# ── Row 23: DELIVERY ─────────────────────────────────────────────────────────
rh(23, 16)
s(23, 2, "DELIVERY", FSB, AL)
input_line(23, 3, 5)
s(23, 6, "%CDC", FSB, AC)
checkbox(23, 7)
input_line(23, 8, 9)
s(23, 10, "%PDC", FSB, AC)
checkbox(23, 11)
s(23, 12, "BEFORE", FSB, AC)
checkbox(23, 13)
# Squeeze AFTER into adjacent columns
m(23, 15, 23, 16)
s(23, 15, "AFTER", FSB, AC)
checkbox(23, 17)

# ── Row 24: SECURITY CHEQUE REQUIREMENT ──────────────────────────────────────
rh(24, 16)
m(24, 2, 24, 4)
s(24, 2, "SECURITY CHEQUE REQUIREMENT", FSB, AL)
s(24, 6, "YES", FSB, AC)
checkbox(24, 7)
s(24, 8, "NO", FSB, AC)
checkbox(24, 9)

# ── Row 25: WHEN WILL IT RECOLLECT ───────────────────────────────────────────
rh(25, 16)
m(25, 2, 25, 4)
s(25, 2, "WHEN WILL IT RECOLLECT", FSB, AL)
input_line(25, 5, 13)

# ── Row 26: WORK IN PROGRESS ─────────────────────────────────────────────────
rh(26, 16)
m(26, 2, 26, 3)
s(26, 2, "WORK IN PROGRESS", FSB, AL)
input_line(26, 4, 11)
s(26, 12, "%", FSB, AC)

# ── Row 27: COMPLETION ───────────────────────────────────────────────────────
rh(27, 16)
s(27, 2, "COMPLETION", FSB, AL)
input_line(27, 3, 5)
s(27, 6, "%CDC", FSB, AC)
checkbox(27, 7)
s(27, 8, "%PDC", FSB, AC)
checkbox(27, 9)

# ── Row 28: TESTING & COMMISSIONING ──────────────────────────────────────────
rh(28, 16)
m(28, 2, 28, 3)
s(28, 2, "TESTING & COMMISSIONING", FSB, AL)
input_line(28, 4, 5)
s(28, 6, "%CDC", FSB, AC)
checkbox(28, 7)
s(28, 8, "%PDC", FSB, AC)
checkbox(28, 9)

# ── Row 29: RETENTION ────────────────────────────────────────────────────────
rh(29, 16)
s(29, 2, "RETENTION", FSB, AL)
input_line(29, 3, 5)
s(29, 6, "%CDC", FSB, AC)
checkbox(29, 7)
s(29, 8, "%PDC", FSB, AC)
checkbox(29, 9)

# ── Row 30: OTHER COMMITTED TERMS ────────────────────────────────────────────
rh(30, 16)
m(30, 2, 30, 4)
s(30, 2, "OTHER COMMITTED TERMS:", FSB, AL)
input_line(30, 5, 19)

# ── Row 31: spacer ────────────────────────────────────────────────────────────
rh(31, 8)

# ── Row 32: ACCOUNTS CONTACT PERSON DETAILS ──────────────────────────────────
rh(32, 16)
section_header(32, "ACCOUNTS CONTACT PERSON DETAILS:")

# ── Row 33: spacer ────────────────────────────────────────────────────────────
rh(33, 6)

# ── Rows 34-36: NAME / EMAIL ID / TEL/MOB ────────────────────────────────────
for r, lbl in [(34, "NAME"), (35, "EMAIL ID"), (36, "TEL/MOB")]:
    rh(r, 16)
    s(r, 2, lbl, FSB, AL)
    input_line(r, 3, 19)

# ── Row 37: spacer ────────────────────────────────────────────────────────────
rh(37, 8)

# ── Row 38: DETAILS OF DOCUMENT HANDOVERING ──────────────────────────────────
rh(38, 16)
section_header(38, "DETAILS OF DOCUMENT HANDOVERING")

# ── Row 39: spacer ────────────────────────────────────────────────────────────
rh(39, 6)

# ── Row 40: 1. INVOICE SUBMISSION ────────────────────────────────────────────
rh(40, 16)
m(40, 2, 40, 5)
s(40, 2, "1. INVOICE SUBMISSION", FSB, AL)
s(40, 8, "OFFICE", FSB, AC)
checkbox(40, 9)
s(40, 12, "SITE", FSB, AC)
checkbox(40, 13)

# ── Row 41: 2. WARRANTY & OPERATIONAL MANUAL SUBMISSION TIME ─────────────────
rh(41, 16)
m(41, 2, 41, 6)
s(41, 2, "2. WARRANTY & OPERATIONAL MANUAL SUBMISSION TIME:", FN, AL)
input_line(41, 7, 19)

# ── Row 42: spacer ────────────────────────────────────────────────────────────
rh(42, 8)

# ── Row 43: PROJECT CONTACT PERSON DETAILS ───────────────────────────────────
rh(43, 16)
section_header(43, "PROJECT CONTACT PERSON DETAILS")

# ── Row 44: spacer ────────────────────────────────────────────────────────────
rh(44, 6)

# ── Rows 45-47: NAME / EMAIL ID / TEL/MOB ────────────────────────────────────
for r, lbl in [(45, "NAME"), (46, "EMAIL ID"), (47, "TEL/MOB")]:
    rh(r, 16)
    s(r, 2, lbl, FSB, AL)
    input_line(r, 3, 19)

# ── Row 48: spacer ────────────────────────────────────────────────────────────
rh(48, 8)

# ── Row 49: ABOVE DETAILS ARE CONFIRMED BY ───────────────────────────────────
rh(49, 14)
full_label(49, "ABOVE DETAILS ARE CONFIRMED BY", FSB)

# ── Row 50: spacer ────────────────────────────────────────────────────────────
rh(50, 6)

# ── Row 51: SALES EXECUTIVE / MANAGER ────────────────────────────────────────
rh(51, 18)
s(51, 2, "SALES EXECUTIVE:", FSB, AL)
input_line(51, 3, 9)
s(51, 15, "MANAGER:", FSB, AL)
input_line(51, 16, 19)

# ── Row 52: SIGNATURE ────────────────────────────────────────────────────────
rh(52, 18)
s(52, 2, "SIGNATURE:", FSB, AL)
input_line(52, 3, 9)
s(52, 15, "SIGNATURE:", FSB, AL)
input_line(52, 16, 19)

# ── Row 53: bottom margin ─────────────────────────────────────────────────────
rh(53, 6)

# ─── Save ──────────────────────────────────────────────────────────────────────
out_path = r"c:\Users\JOYAL'S LEGION\OneDrive\Documents\GRP\GRP_SYS\Order_Confirmation_Form.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
