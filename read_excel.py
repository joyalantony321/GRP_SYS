import openpyxl
wb = openpyxl.load_workbook(r"c:\Users\JOYAL'S LEGION\OneDrive\Documents\GRP\GRP_SYS\client\Public\order conformation Forms\ORDER CONFIRMATION EXCEL.xlsx")
print('Sheets:', wb.sheetnames)
ws = wb.active
print('Max row:', ws.max_row, 'Max col:', ws.max_column)
print('Merged cells:', list(ws.merged_cells.ranges)[:20])
count = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            print(f'{cell.coordinate}: {repr(cell.value)}')
            count += 1
print(f'Total non-empty: {count}')

# Check column widths and row heights
print('Col widths:', {k: v.width for k, v in list(ws.column_dimensions.items())[:10]})
print('Row heights:', {k: v.height for k, v in list(ws.row_dimensions.items())[:10]})
